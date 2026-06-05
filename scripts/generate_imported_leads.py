from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, time, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FILES = [
    Path("/Users/kacperbernecki/Downloads/Leady WSZYSTKIE 2026.xlsx"),
    Path("/Users/kacperbernecki/Downloads/Re energy leady.xlsx"),
]
OUTPUT_FILE = PROJECT_ROOT / "lib" / "imported-leads.ts"
LEADS_ALL_YEAR = 2026
RE_ENERGY_YEAR = 2025
FALLBACK_CREATED_AT = f"{LEADS_ALL_YEAR}-01-01T12:00:00Z"
FALLBACK_CREATED_AT_BY_YEAR = {
    LEADS_ALL_YEAR: f"{LEADS_ALL_YEAR}-01-01T12:00:00Z",
    RE_ENERGY_YEAR: f"{RE_ENERGY_YEAR}-01-01T12:00:00Z",
}

VOIVODESHIPS = {
    "dolnośląskie",
    "kujawsko-pomorskie",
    "lubelskie",
    "lubuskie",
    "łódzkie",
    "małopolskie",
    "mazowieckie",
    "opolskie",
    "podkarpackie",
    "podlaskie",
    "pomorskie",
    "śląskie",
    "świętokrzyskie",
    "warmińsko-mazurskie",
    "wielkopolskie",
    "zachodniopomorskie",
}

VOIVODESHIP_ALIASES = {
    "podkarpacie": "Podkarpackie",
    "mazowsze": "Mazowieckie",
    "busko": "Świętokrzyskie",
    "grzymala": "Świętokrzyskie",
    "marchocice": "Świętokrzyskie",
    "polaniec": "Świętokrzyskie",
    "sandomierz": "Świętokrzyskie",
    "skarzysko": "Świętokrzyskie",
    "stojewsko": "Świętokrzyskie",
    "wloszczowa": "Świętokrzyskie",
    "cierpisz": "Podkarpackie",
    "przeworsk": "Podkarpackie",
    "38-232": "Podkarpackie",
    "37-124": "Podkarpackie",
    "hedwizyn": "Lubelskie",
    "lublin": "Lubelskie",
    "markuszow": "Lubelskie",
}


@dataclass
class ImportRow:
    name: str
    phone: str
    phone_key: str
    email: str | None
    email_key: str | None
    created_at: str | None
    address: str | None
    voivodeship: str | None
    postal_code: str | None
    source: str
    status: str
    comment: str | None
    row_label: str
    source_group: str


@dataclass
class LeadAggregate:
    id: str
    name: str
    phone: str
    phone_key: str | None
    email: str | None
    email_key: str | None
    created_at: str
    updated_at: str
    address: str | None
    voivodeship: str | None
    postal_code: str | None
    source: str
    source_group: str
    status: str
    comments: list[tuple[str, str, str]] = field(default_factory=list)
    seen_comments: set[str] = field(default_factory=set)


def text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value).replace("\u00a0", " ").strip()


def clean_text(value: Any, limit: int = 500) -> str:
    return re.sub(r"\s+", " ", text(value)).strip()[:limit]


def normalize_key(value: str) -> str:
    replacements = str.maketrans("ąćęłńóśżźĄĆĘŁŃÓŚŻŹ", "acelnoszzACELNOSZZ")
    return value.translate(replacements).lower().replace("_", " ").strip()


def normalize_phone(value: Any) -> tuple[str, str] | tuple[None, None]:
    raw = text(value)
    digits = re.sub(r"\D", "", raw)
    if len(digits) < 9:
        return None, None
    last9 = digits[-9:]
    if not re.match(r"^[4-9]\d{8}$", last9):
        return None, None
    dial = f"+48{last9}"
    return f"+48 {last9[:3]} {last9[3:6]} {last9[6:]}", f"phone:{dial}"


def email_key(value: Any) -> tuple[str | None, str | None]:
    value_text = clean_text(value, 180).lower()
    if "@" not in value_text:
        return None, None
    match = re.search(r"[\w.+-]+@[\w.-]+\.[a-z]{2,}", value_text)
    if not match:
        return None, None
    email = match.group(0)
    return email, f"email:{email}"


def normalize_year(year: str, force_year: int | None = None) -> int:
    if force_year:
        return force_year
    if len(year) == 2:
        return 2000 + int(year)
    return int(year)


def replace_year(value: datetime, force_year: int | None) -> datetime:
    if not force_year:
        return value
    try:
        return value.replace(year=force_year)
    except ValueError:
        return value.replace(year=force_year, day=28)


def iso_datetime(value: datetime, force_year: int | None = None) -> str:
    value = replace_year(value, force_year)
    if value.tzinfo:
        return value.astimezone(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    return value.isoformat(timespec="seconds") + "Z"


def parse_numeric_date(value_text: str, force_year: int | None = None) -> str | None:
    match = re.match(r"^(\d{1,2})([./])(\d{1,2})(?:[./](\d{2,4}))?(?=$|\s|[,;:)])", value_text)
    if not match:
        return None

    first = int(match.group(1))
    separator = match.group(2)
    second = int(match.group(3))
    year_text = match.group(4)
    if not year_text and not force_year:
        return None

    year = normalize_year(year_text or str(force_year), force_year)

    if separator == "/" and second > 12:
        day = second
        month = first
    else:
        day = first
        month = second

    try:
        return datetime(year, month, day, 12).isoformat(timespec="seconds") + "Z"
    except ValueError:
        return None


def parse_date(value: Any, force_year: int | None = None) -> str | None:
    if isinstance(value, datetime):
        return iso_datetime(value, force_year)
    if isinstance(value, date):
        return iso_datetime(datetime.combine(value, time(hour=12)), force_year)

    value_text = clean_text(value, 80)
    if not value_text:
        return None

    if re.match(r"^\d{4}-\d{2}-\d{2}([T\s]|$)", value_text):
        try:
            parsed = datetime.fromisoformat(value_text.replace("Z", "+00:00"))
            return iso_datetime(parsed, force_year)
        except ValueError:
            pass

    numeric = parse_numeric_date(value_text, force_year)
    if numeric:
        return numeric

    patterns = (
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
    )
    for candidate in dict.fromkeys([value_text, value_text[:10]]):
        for pattern in patterns:
            try:
                parsed = datetime.strptime(candidate, pattern)
                return iso_datetime(parsed, force_year)
            except ValueError:
                pass
    return None


def is_probable_date_cell(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    value_text = clean_text(value, 80)
    return bool(
        re.match(r"^\d{4}-\d{2}-\d{2}([T\s]|$)", value_text)
        or re.match(r"^\d{1,2}[./]\d{1,2}[./]\d{2,4}(\s|$)", value_text)
    )


def normalized_name(value: Any) -> str:
    value_text = clean_text(value, 180)
    value_text = re.sub(r"[\w.+-]+@[\w.-]+\.[a-z]{2,}", " ", value_text, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", value_text).strip()


def looks_like_person_name(value: str) -> bool:
    normalized = normalize_key(value)
    if not value or normalized in {"imie", "imie i nazwisko", "telefon", "adres", "brak"}:
        return False
    if is_probable_date_cell(value):
        return False
    if re.search(r"\d", value):
        return False
    return bool(re.search(r"[A-Za-zĄĆĘŁŃÓŚŻŹąćęłńóśżź]{2,}", value))


def name_from_row(row: tuple[Any, ...], phone_index: int) -> str:
    preferred_indexes: list[int] = []
    if phone_index >= 3:
        preferred_indexes.append(2)
    if phone_index > 0:
        preferred_indexes.append(phone_index - 1)
    preferred_indexes.extend(range(max(0, phone_index - 3), phone_index))

    for index in dict.fromkeys(preferred_indexes):
        if index < len(row):
            name = normalized_name(row[index])
            if looks_like_person_name(name):
                return name
    return ""


def comment_date_from_text(value: Any, force_year: int | None = None) -> str | None:
    value_text = clean_text(value, 300)
    if not value_text:
        return None
    for pattern in (
        r"\b\d{4}-\d{2}-\d{2}(?:[T\s]\d{2}:\d{2}(?::\d{2})?(?:\.\d+)?(?:Z|[+-]\d{2}:?\d{2})?)?\b",
        r"\b\d{1,2}\.\d{1,2}\.\d{4}\b",
        r"\b\d{1,2}\.\d{1,2}\.\d{2}\b",
        r"\b\d{1,2}\.\d{1,2}\b",
        r"\b\d{1,2}/\d{1,2}/\d{4}\b",
        r"\b\d{1,2}/\d{1,2}/\d{2}\b",
        r"\b\d{1,2}/\d{1,2}(?![A-Za-zĄĆĘŁŃÓŚŻŹąćęłńóśżź])\b",
    ):
        match = re.search(pattern, value_text)
        if match:
            parsed = parse_date(match.group(0), force_year)
            if parsed:
                return parsed
    return None


def row_date_from_cells(row: tuple[Any, ...], force_year: int | None = None) -> str | None:
    for value in row:
        parsed = parse_date(value, force_year)
        if parsed:
            return parsed
    return None


def postal_code_from(*values: Any) -> str | None:
    joined = " ".join(clean_text(value, 120) for value in values)
    match = re.search(r"\b\d{2}-\d{3}\b", joined)
    return match.group(0) if match else None


def voivodeship_from(*values: Any) -> str | None:
    joined = " ".join(clean_text(value, 120).lower() for value in values)
    normalized_joined = normalize_key(joined)
    for alias, voivodeship in VOIVODESHIP_ALIASES.items():
        if alias in normalized_joined:
            return voivodeship
    for item in VOIVODESHIPS:
        if item in joined:
            return item.capitalize()
    return None


def map_status(status: str | None, comment: str | None, source: str, created_at: str | None) -> str:
    context = " ".join(part for part in [status or "", comment or "", source] if part).lower()
    rules = [
        (r"um[oó]w|spotkan|kalendarz|wizyta", "Spotkanie"),
        (r"callback|call back|oddzwo|ponowny kontakt", "Call back"),
        (r"\bn\.?\s*o\.?\b|nie\s*odbiera|nieodebra|nie odebra|brak kontaktu", "Nie odebrał"),
        (r"pomy[lł]ka numeru|nie ma takiego numeru|s[lł]u[zż]bowy numer|b[lł][eę]dny|z[lł]y numer|niepoprawny numer", "Błędny numer"),
        (r"rezygn|nie chce|brak zgody|ju[zż] ma|nie potrzebuje|nie op[lł]aca|nic nie dop[lł]aca|nie aktualne", "Rezygnacja"),
        (r"zwrot", "Zwrot"),
        (r"weryfik", "Do weryfikacji"),
        (r"umowa|podpis", "Umowa"),
        (r"po spotk", "Po spotkaniu"),
        (r"przypis", "Przypisany"),
        (r"created|utworzon|nie dzwonione|niedzwonione|now", "Nowy"),
        (r"zimn|stara baza|cold", "Zimna baza"),
    ]
    for pattern, mapped in rules:
        if re.search(pattern, context, re.IGNORECASE):
            return mapped
    return "Nowy" if created_at and created_at.startswith("2026") else "Zimna baza"


def status_score(status: str) -> int:
    order = [
        "Zimna baza",
        "Nowy",
        "Nie odebrał",
        "Call back",
        "Do weryfikacji",
        "Przypisany",
        "Spotkanie",
        "Po spotkaniu",
        "Oferta wysłana",
        "Oferta zaakceptowana",
        "Podpis elektroniczny",
        "Umowa",
        "W realizacji",
        "Zrealizowany",
    ]
    if status in ["Błędny numer", "Rezygnacja", "Zwrot"]:
        return 50
    return order.index(status) if status in order else 0


def find_phone_index(row: tuple[Any, ...]) -> tuple[int, str, str] | None:
    for index, value in enumerate(row):
        if is_probable_date_cell(value):
            continue
        normalized = normalize_phone(value)
        if normalized[0] and normalized[1]:
            return index, normalized[0], normalized[1]
    return None


def first_email(row: tuple[Any, ...]) -> tuple[str | None, str | None]:
    for value in row:
        email, key = email_key(value)
        if email and key:
            return email, key
    return None, None


def build_comment(source: str, row_label: str, *parts: Any) -> str | None:
    values = [clean_text(part, 400) for part in parts if clean_text(part, 400)]
    values = [value for value in values if value.lower() not in {"brak", "-", "x"}]
    if not values:
        return None
    return f"{source}, {row_label}: " + " | ".join(dict.fromkeys(values))


def meta_source(file_name: str, sheet_name: str) -> str:
    return f"{file_name} / {sheet_name}"


def rows_from_meta_workbook(path: Path) -> list[ImportRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[ImportRow] = []

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        headers = [normalize_key(clean_text(cell)) for cell in next(rows, [])]
        header_map = {header: index for index, header in enumerate(headers) if header}
        source = meta_source(path.name, sheet.title)

        def get(row: tuple[Any, ...], *names: str) -> Any:
            for name in names:
                index = header_map.get(normalize_key(name))
                if index is not None and index < len(row):
                    return row[index]
            return None

        for row_number, row in enumerate(rows, start=2):
            phone, phone_key = normalize_phone(get(row, "phone_number", "phone", "telefon"))
            email, key = email_key(get(row, "email", "e-mail"))
            if not phone_key and not key:
                continue

            name = normalized_name(get(row, "full_name", "name", "imie i nazwisko")) or f"Klient {phone or email}"
            created_at = (
                parse_date(get(row, "created_time", "created_at", "data"), LEADS_ALL_YEAR)
                or FALLBACK_CREATED_AT_BY_YEAR[LEADS_ALL_YEAR]
            )
            row_source = clean_text(get(row, "form_name", "campaign_name", "adset_name"), 160) or source
            status_text = clean_text(get(row, "lead_status", "status"), 80)
            comment = build_comment(source, f"wiersz {row_number}", get(row, "comment", "komentarz", "uwagi"))
            address = clean_text(get(row, "adres", "address"), 250) or None
            postal = clean_text(get(row, "post_code", "postal_code", "kod pocztowy"), 20) or postal_code_from(row)
            voivodeship = clean_text(get(row, "województwo", "wojewodztwo", "voivodeship"), 80) or voivodeship_from(sheet.title, row)

            output.append(
                ImportRow(
                    name=name,
                    phone=phone or "",
                    phone_key=phone_key or key or "",
                    email=email,
                    email_key=key,
                    created_at=created_at,
                    address=address,
                    voivodeship=voivodeship,
                    postal_code=postal,
                    source=row_source,
                    status=map_status(status_text, comment, source, created_at),
                    comment=comment,
                    row_label=f"{sheet.title} #{row_number}",
                    source_group="leady_wszystkie_2026",
                )
            )

    return output


def rows_from_re_energy(path: Path) -> list[ImportRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: list[ImportRow] = []

    for sheet in workbook.worksheets:
        source = meta_source(path.name, sheet.title)
        sheet_key = normalize_key(sheet.title)

        if "kalendarz" in sheet_key:
            output.extend(rows_from_calendar_sheet(path.name, sheet))
            continue

        for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            phone_info = find_phone_index(row)
            if not phone_info:
                continue

            phone_index, phone, phone_key = phone_info
            email, key = first_email(row)
            created_at = parse_date(row[0] if row else None, RE_ENERGY_YEAR)
            name = name_from_row(row, phone_index)
            if not name:
                name = f"Klient {phone}"

            right = list(row[phone_index + 1 :])
            address_parts = [value for value in right[:2] if clean_text(value, 140) and not voivodeship_from(value)]
            address = ", ".join(clean_text(value, 140) for value in address_parts) or None
            voivodeship = voivodeship_from(sheet.title, *right)
            postal = postal_code_from(*right)

            product = right[2] if len(right) > 2 else None
            comment_cell = right[3] if len(right) > 3 else None
            if "podkarpacie" in sheet_key and len(right) > 4:
                comment_cell = right[4]
            if "styczen" in sheet_key or "luty" in sheet_key:
                product = right[1] if len(right) > 1 else product
                comment_cell = right[2] if len(right) > 2 else comment_cell
            if not created_at:
                created_at = comment_date_from_text(comment_cell, RE_ENERGY_YEAR)
            if not created_at:
                created_at = row_date_from_cells(row, RE_ENERGY_YEAR)
            if not created_at:
                created_at = FALLBACK_CREATED_AT_BY_YEAR[RE_ENERGY_YEAR]

            comment = build_comment(source, f"wiersz {row_number}", product, comment_cell)
            status_hint = "Spotkanie" if "spotkania" in sheet_key else clean_text(row[1] if len(row) > 1 else "", 80)

            output.append(
                ImportRow(
                    name=name,
                    phone=phone,
                    phone_key=phone_key,
                    email=email,
                    email_key=key,
                    created_at=created_at,
                    address=address,
                    voivodeship=voivodeship,
                    postal_code=postal,
                    source=source,
                    status=map_status(status_hint, comment, source, created_at),
                    comment=comment,
                    row_label=f"{sheet.title} #{row_number}",
                    source_group="re_energy_2025",
                )
            )

    return output


def rows_from_calendar_sheet(file_name: str, sheet: Any) -> list[ImportRow]:
    output: list[ImportRow] = []
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 3:
        return output

    days = rows[0]
    for row_number, row in enumerate(rows[2:], start=3):
        for start in range(1, len(row), 4):
            phone_cell = row[start + 2] if start + 2 < len(row) else None
            normalized = normalize_phone(phone_cell)
            if not normalized[0] or not normalized[1]:
                continue
            name = clean_text(row[start + 1] if start + 1 < len(row) else "", 180) or f"Klient {normalized[0]}"
            hour = clean_text(row[start] if start < len(row) else "", 80)
            address = clean_text(row[start + 3] if start + 3 < len(row) else "", 220) or None
            day = clean_text(days[start] if start < len(days) else "", 80)
            source = meta_source(file_name, sheet.title)
            comment = build_comment(source, f"wiersz {row_number}", f"Termin: {day} {hour}".strip(), address)
            output.append(
                ImportRow(
                    name=name,
                    phone=normalized[0],
                    phone_key=normalized[1],
                    email=None,
                    email_key=None,
                    created_at=FALLBACK_CREATED_AT_BY_YEAR[RE_ENERGY_YEAR],
                    address=address,
                    voivodeship=voivodeship_from(address),
                    postal_code=postal_code_from(address),
                    source=source,
                    status="Spotkanie",
                    comment=comment,
                    row_label=f"{sheet.title} #{row_number}",
                    source_group="re_energy_2025",
                )
            )
    return output


def add_comment(lead: LeadAggregate, source: str, row_label: str, comment: str | None) -> None:
    if not comment:
        return
    key = normalize_key(comment)
    if key in lead.seen_comments:
        return
    lead.seen_comments.add(key)
    lead.comments.append((source, row_label, comment))


def aggregate(rows: list[ImportRow]) -> tuple[list[LeadAggregate], int]:
    leads: dict[str, LeadAggregate] = {}
    duplicate_rows = 0

    for index, row in enumerate(rows, start=1):
        key = row.phone_key or row.email_key
        if not key:
            continue

        lead_id = "imported-lead-" + re.sub(r"[^a-z0-9]+", "-", key.lower()).strip("-")
        created_at = row.created_at or FALLBACK_CREATED_AT

        if key not in leads:
            leads[key] = LeadAggregate(
                id=lead_id,
                name=row.name,
                phone=row.phone,
                phone_key=row.phone_key,
                email=row.email,
                email_key=row.email_key,
                created_at=created_at,
                updated_at=created_at,
                address=row.address,
                voivodeship=row.voivodeship,
                postal_code=row.postal_code,
                source=row.source,
                source_group=row.source_group,
                status=row.status,
            )
        else:
            duplicate_rows += 1
            lead = leads[key]
            lead.email = lead.email or row.email
            lead.email_key = lead.email_key or row.email_key
            lead.address = lead.address or row.address
            lead.voivodeship = lead.voivodeship or row.voivodeship
            lead.postal_code = lead.postal_code or row.postal_code
            if row.created_at and row.source_group == lead.source_group:
                if lead.created_at == FALLBACK_CREATED_AT or row.created_at < lead.created_at:
                    lead.created_at = row.created_at
                if lead.updated_at == FALLBACK_CREATED_AT:
                    lead.updated_at = row.created_at
                else:
                    lead.updated_at = max(lead.updated_at, row.created_at)
            if status_score(row.status) > status_score(lead.status):
                lead.status = row.status

        add_comment(leads[key], row.source, row.row_label, row.comment)

    return list(leads.values()), duplicate_rows


def lead_to_dict(lead: LeadAggregate) -> dict[str, Any]:
    return {
        "id": lead.id,
        "full_name": lead.name,
        "postal_code": lead.postal_code,
        "phone": lead.phone,
        "email": lead.email,
        "phone_key": lead.phone_key,
        "email_key": lead.email_key,
        "address": lead.address,
        "voivodeship": lead.voivodeship,
        "county": None,
        "status": lead.status,
        "assigned_to": None,
        "created_at": lead.created_at,
        "updated_at": lead.updated_at,
        "last_opened_at": None,
        "source": lead.source,
        "resignation_reason": None,
        "callback_at": None,
        "meeting_at": None,
        "meeting_address": None,
        "meeting_note": None,
        "contract_number": None,
        "crm_environment": "production",
    }


def history_to_dict(lead: LeadAggregate, index: int, comment: tuple[str, str, str]) -> dict[str, Any]:
    source, row_label, description = comment
    return {
        "id": f"{lead.id}-comment-{index + 1}",
        "lead_id": lead.id,
        "user_id": "kacper-admin",
        "action_type": "import_comment",
        "description": description,
        "old_value": None,
        "new_value": {"source": source, "row": row_label},
        "created_at": lead.created_at,
    }


def write_output(leads: list[LeadAggregate], duplicate_rows: int, files: list[Path], source_rows: int) -> None:
    lead_rows = [lead_to_dict(lead) for lead in leads]
    histories = [history_to_dict(lead, index, comment) for lead in leads for index, comment in enumerate(lead.comments)]
    generated_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")
    version = f"imported-leads-{generated_at}-{len(lead_rows)}"
    payload = (
        'import type { Lead, LeadHistory } from "@/lib/types";\n\n'
        "export const importedLeadSeedMeta = "
        + json.dumps(
            {
                "version": version,
                "generatedAt": generated_at,
                "sourceFiles": [path.name for path in files],
                "sourceRows": source_rows,
                "uniqueLeads": len(lead_rows),
                "duplicateRows": duplicate_rows,
                "commentRows": len(histories),
            },
            ensure_ascii=False,
            indent=2,
        )
        + ";\n\n"
        + "export const importedLeadSeedRows = "
        + json.dumps(lead_rows, ensure_ascii=False, indent=2)
        + " satisfies Lead[];\n\n"
        + "export const importedLeadHistorySeedRows = "
        + json.dumps(histories, ensure_ascii=False, indent=2)
        + " satisfies LeadHistory[];\n"
    )
    OUTPUT_FILE.write_text(payload, encoding="utf-8")


def main() -> int:
    files = [Path(arg) for arg in sys.argv[1:]] or DEFAULT_FILES
    missing = [path for path in files if not path.exists()]
    if missing:
        print("Brakuje plików:")
        for path in missing:
            print(f"- {path}")
        return 2

    rows: list[ImportRow] = []
    for path in files:
        if path.name.lower().startswith("leady wszystkie"):
            rows.extend(rows_from_meta_workbook(path))
        else:
            rows.extend(rows_from_re_energy(path))

    leads, duplicate_rows = aggregate(rows)
    write_output(leads, duplicate_rows, files, len(rows))
    print(f"Gotowe: {len(leads)} leadów, {duplicate_rows} duplikatów, {sum(len(lead.comments) for lead in leads)} komentarzy.")
    print(f"Zapisano: {OUTPUT_FILE}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
